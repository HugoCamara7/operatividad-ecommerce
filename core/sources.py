# -*- coding: utf-8 -*-
"""Fuentes de datos.

La aplicación consume dos operaciones: `headers()` para descubrir qué hay en el
origen (barato) y `load()` para traer una tabla concreta (caro).  Separarlas
permite decidir qué extraer antes de pagar el costo, y hace que sustituir el
Excel por una base de datos no afecte a nada aguas abajo.
"""
from __future__ import annotations

import hashlib
import io
import re
import zipfile
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree.ElementTree import ParseError

import pandas as pd

from . import pivotcache


@dataclass
class TableHeader:
    """Descripción barata de una tabla disponible en el origen."""

    name: str
    columns: list[str]
    origin: str = ""
    cost: int = 0            # tamaño aproximado; se usa para desempatar candidatos
    rows_hint: int = 0


class DataSource(ABC):
    @abstractmethod
    def headers(self) -> list[TableHeader]:
        """Tablas disponibles y sus columnas, sin leer los datos."""

    @abstractmethod
    def load(self, name: str) -> pd.DataFrame:
        """Trae una tabla completa, ya identificada por `headers()`."""

    @property
    @abstractmethod
    def fingerprint(self) -> str:
        """Identificador estable del contenido, usado para cachear."""

    @property
    def label(self) -> str:
        return self.__class__.__name__


# ---------------------------------------------------------------------------
#  Excel
# ---------------------------------------------------------------------------
class ExcelSource(DataSource):
    """Lee un .xlsx.

    El libro de Operatividad guarda sus tablas dinámicas sobre hojas externas: en
    el archivo sólo quedan los resultados visibles y el *pivot cache*, que sí trae
    el detalle registro a registro.  Por eso se leen primero los caches y las
    hojas planas quedan como respaldo.
    """

    #: Una hoja necesita al menos una fila de datos bajo el encabezado. No se
    #: filtra por volumen: las hojas de resumen se descartan solas porque no
    #: contienen las columnas firma de ningún conjunto.
    MIN_FLAT_ROWS = 2

    def __init__(self, data: bytes | str | Path, name: str = ""):
        if isinstance(data, (str, Path)):
            path = Path(data)
            self._bytes = path.read_bytes()
            self._name = name or path.name
        else:
            self._bytes = bytes(data)
            self._name = name or "archivo.xlsx"
        self._fingerprint = ""
        self._headers: list[TableHeader] | None = None
        self._cache_fields: dict[int, tuple] = {}

    @property
    def label(self) -> str:
        return self._name

    @property
    def fingerprint(self) -> str:
        if not self._fingerprint:
            digest = hashlib.sha1(self._bytes).hexdigest()[:16]
            self._fingerprint = f"{len(self._bytes)}-{digest}"
        return self._fingerprint

    def _zip(self) -> zipfile.ZipFile:
        return zipfile.ZipFile(io.BytesIO(self._bytes))

    # -- descubrimiento -----------------------------------------------------
    def headers(self) -> list[TableHeader]:
        if self._headers is not None:
            return self._headers
        out: list[TableHeader] = []
        with self._zip() as zf:
            sizes = {i.filename: i.file_size for i in zf.infolist()}
            for index in pivotcache.list_caches(zf):
                try:
                    names, shared = pivotcache.cache_fields(zf, index)
                except (KeyError, ParseError):
                    continue
                if not names:
                    continue
                self._cache_fields[index] = (names, shared)
                out.append(
                    TableHeader(
                        name=f"pivotcache{index}",
                        columns=_dedupe(names),
                        origin=f"pivot cache #{index}",
                        cost=sizes.get(f"xl/pivotCache/pivotCacheRecords{index}.xml", 0),
                    )
                )
            out.extend(self._flat_headers(zf))
        self._headers = out
        return out

    def _flat_headers(self, zf: zipfile.ZipFile) -> list[TableHeader]:
        """Encabezados de las hojas normales, leídos en modo sólo-lectura."""
        out: list[TableHeader] = []
        try:
            import openpyxl

            book = openpyxl.load_workbook(io.BytesIO(self._bytes), read_only=True, data_only=True)
        except Exception:
            return out
        for sheet in book.worksheets:
            if sheet.max_row is None or sheet.max_row < self.MIN_FLAT_ROWS:
                continue
            try:
                first = next(sheet.iter_rows(values_only=True), None)
            except Exception:
                continue
            if not first:
                continue
            columns = [str(c) for c in first if c is not None]
            if len(columns) < 3:
                continue
            out.append(
                TableHeader(
                    name=f"hoja::{sheet.title}",
                    columns=_dedupe(columns),
                    origin=f"hoja '{sheet.title}'",
                    cost=10**9,          # se prefiere siempre un pivot cache
                    rows_hint=sheet.max_row,
                )
            )
        book.close()
        return out

    # -- carga --------------------------------------------------------------
    def load(self, name: str) -> pd.DataFrame:
        if name.startswith("pivotcache"):
            index = int(name.removeprefix("pivotcache"))
            with self._zip() as zf:
                names, shared = self._cache_fields.get(index) or pivotcache.cache_fields(zf, index)
                rows = pivotcache.read_records(zf, index, names, shared)
            return pd.DataFrame(rows, columns=_dedupe(names))
        if name.startswith("hoja::"):
            sheet = name.removeprefix("hoja::")
            frame = pd.read_excel(io.BytesIO(self._bytes), sheet_name=sheet, engine="openpyxl")
            frame = frame.dropna(axis=1, how="all").dropna(axis=0, how="all")
            frame.columns = _dedupe([str(c) for c in frame.columns])
            return frame
        raise KeyError(f"Tabla desconocida: {name}")


def _dedupe(names: list[str]) -> list[str]:
    """Evita nombres de columna repetidos conservando el primero."""
    seen: dict[str, int] = {}
    out = []
    for raw in names:
        name = "" if raw is None else str(raw)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


# ---------------------------------------------------------------------------
#  Plantilla para el futuro origen directo (base de datos / API)
# ---------------------------------------------------------------------------
class SQLSource(DataSource):
    """Esqueleto para conectar el dashboard a una base de datos.

    Al existir la conexión directa basta con completar las consultas: el resto de
    la aplicación (normalización, KPIs, gráficos, interfaz) no cambia, porque
    consume exactamente el mismo contrato.
    """

    def __init__(self, connection, queries: dict[str, str]):
        self._connection = connection
        self._queries = queries

    @property
    def fingerprint(self) -> str:
        payload = repr(sorted(self._queries.items())).encode()
        return "sql-" + hashlib.sha1(payload).hexdigest()[:16]

    def headers(self) -> list[TableHeader]:
        out = []
        for name, query in self._queries.items():
            sample = pd.read_sql(f"SELECT * FROM ({query}) t LIMIT 1", self._connection)
            out.append(TableHeader(name=name, columns=list(sample.columns), origin="sql"))
        return out

    def load(self, name: str) -> pd.DataFrame:
        return pd.read_sql(self._queries[name], self._connection)
